import pandas as pd
import re
import os
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def deduplicate_columns(columns):
    seen = {}
    new_columns = []
    for col in columns:
        if col in seen:
            seen[col] += 1
            new_columns.append(f"{col}.{seen[col]}")  # Append a suffix to duplicate column names
        else:
            seen[col] = 0
            new_columns.append(col)
    return new_columns

def convert_text_to_excel(file_path, output_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    header_match = re.search(r'DATE:- (.*?)C\.B\.S\.E\. - (.*?)REGION: (.*?)PAGE:-', content)
    school_match = re.search(r'SCHOOL : - (\d+)\s+([^\n]+)', content)

    school_info = ""
    if header_match:
        date, exam_info, region = header_match.groups()
        school_info += f"Date: {date.strip()} | Exam: {exam_info.strip()} | Region: {region.strip()}\n"
    if school_match:
        school_code, school_name = school_match.groups()
        school_info += f"School Code: {school_code} | School Name: {school_name.strip()}"

    students = []
    records = re.finditer(
        r'(\d{8})\s+([MF])\s+([^\n]+?)\s+((?:\d{3}\s+)+)\s+(PASS|FAIL)\s*'
        r'(.*?)(?:\n|$)',
        content,
        re.DOTALL
    )

    for record in records:
        roll_no, gender, name, subject_codes, result, grades_str = record.groups()
        name = ' '.join(name.split())  # Clean up name
        subject_codes = subject_codes.strip().split()
        grade_pairs = re.findall(r'(\d{3})\s+([A-Z]\d)', grades_str)

        student_data = {
            'Roll No': roll_no,
            'Name': name,
            'Gender': gender,
            'Result': result
        }

        for i, (code, grade_pair) in enumerate(zip(subject_codes, grade_pairs)):
            if i < len(grade_pairs):
                marks, grade = grade_pair
                student_data[f'Sub {code} Marks'] = marks
                student_data[f'Sub {code} Grade'] = grade

        students.append(student_data)

    if not students:
        raise ValueError("No valid student records found!")

    df = pd.DataFrame(students)

    base_columns = ['Roll No', 'Name', 'Gender']
    subject_columns = []

    for col in df.columns:
        if col.startswith('Sub') and 'Marks' in col:
            sub_code = col.split()[1]
            if f'Sub {sub_code} Marks' not in subject_columns and f'Sub {sub_code} Grade' not in subject_columns:
                subject_columns.extend([f'Sub {sub_code} Marks', f'Sub {sub_code} Grade'])

    columns_order = base_columns + subject_columns + ['Result']
    df = df[columns_order]

    df = df.drop_duplicates(keep='first')
    df.columns = deduplicate_columns(df.columns)

    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    df.to_excel(writer, index=False, startrow=4, header=True)

    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    worksheet.merge_cells('A1:N1')
    worksheet.merge_cells('A2:N2')

    header_cell = worksheet['A1']
    header_cell.value = school_info.split('\n')[0]
    header_cell.font = Font(bold=True)
    header_cell.alignment = Alignment(horizontal='center')

    school_cell = worksheet['A2']
    school_cell.value = school_info.split('\n')[1] if '\n' in school_info else ""
    school_cell.font = Font(bold=True)
    school_cell.alignment = Alignment(horizontal='center')

    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=5, column=col_num)
        cell.font = Font(bold=True)

    worksheet.freeze_panes = 'A6'
    for i, column_cells in enumerate(df.columns, 1):
        max_length = max(
            [len(str(col))] + [len(str(cell)) for cell in df[col].astype(str)]
        )
        col_letter = get_column_letter(i)
        adjusted_width = 12 if col_letter == 'A' else max_length + 2
        worksheet.column_dimensions[col_letter].width = adjusted_width

    writer.close()
